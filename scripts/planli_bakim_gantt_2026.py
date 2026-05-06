#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
2026 Yılı Planlı Bakım Gantt Takvimi - Excel Çıktısı
Son bakım tarihlerinden yola çıkarak 16 Nisan 2026 – 31 Aralık 2026 arasındaki
gelecek bakım tarihlerini hesaplar ve Gantt şeklinde Excel'e yazar.
"""

import os
from collections import defaultdict
from datetime import date, timedelta
from dateutil.relativedelta import relativedelta
from calendar import monthrange

import mysql.connector
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ─── Ayarlar ────────────────────────────────────────────────────────────────
DB_CONFIG = {
    "host": "127.0.0.1",
    "user": "root",
    "password": "",
    "database": "parkbkm",
    "charset": "utf8mb4",
}

BUGUN = date(2026, 4, 16)
YIL_SONU = date(2026, 12, 31)

AY_ISIMLERI = [
    "", "Ocak", "Şubat", "Mart", "Nisan", "Mayıs", "Haziran",
    "Temmuz", "Ağustos", "Eylül", "Ekim", "Kasım", "Aralık",
]

# Periyot -> ay cinsinden süre
PERIYOT_AY = {
    "Periyodik: 1 Ay": 1,
    "Periyodik: 3 Ay": 3,
    "Periyodik: 6 Ay": 6,
    "Periyodik: 1 Yıl": 12,
}

# Renkler (hex) - periyot bazlı
PERIYOT_RENK = {
    1:  "3498DB",  # Mavi - 1 Ay
    3:  "2ECC71",  # Yeşil - 3 Ay
    6:  "F39C12",  # Turuncu - 6 Ay
    12: "E74C3C",  # Kırmızı - 1 Yıl
}

PERIYOT_RENK_ACIK = {
    1:  "D6EAF8",
    3:  "D5F5E3",
    6:  "FDEBD0",
    12: "FADBD8",
}

OUTPUT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Planli_Bakim_Gantt_2026.xlsx")


# ─── Veritabanı ─────────────────────────────────────────────────────────────
def fetch_last_maintenance():
    """Her ekipman+tanım+periyot grubu için son bakım tarihini çeker."""
    conn = mysql.connector.connect(**DB_CONFIG)
    cursor = conn.cursor(dictionary=True)
    cursor.execute("""
        SELECT kodu, tanimi, periyodu, MAX(tarihi) as son_bakim
        FROM planlibakim
        WHERE durumu != 'ötelendi'
        GROUP BY kodu, tanimi, periyodu
        ORDER BY kodu, periyodu
    """)
    rows = cursor.fetchall()
    cursor.close()
    conn.close()
    return rows


# ─── Gelecek Bakım Hesaplama ────────────────────────────────────────────────
def parse_periyot(periyodu_str):
    """Periyot string'inden ay sayısını döndürür."""
    for key, ay in PERIYOT_AY.items():
        if key in periyodu_str:
            return ay
    # Fallback: sayıyı parse etmeye çalış
    for token in periyodu_str.split():
        if token.isdigit():
            return int(token)
    return None


def hesapla_gelecek_bakimlar(rows):
    """
    Her grup için son bakım tarihinden periyot kadar ilerleyerek
    BUGUN ile YIL_SONU arasındaki tüm gelecek bakım tarihlerini üretir.
    """
    sonuc = []  # (kodu, tanimi, periyot_ay, [tarih1, tarih2, ...])

    for r in rows:
        kodu = r["kodu"] or ""
        tanimi = r["tanimi"] or ""
        periyodu = r["periyodu"] or ""
        son_bakim = r["son_bakim"]

        if not son_bakim:
            continue

        if isinstance(son_bakim, str):
            from datetime import datetime
            son_bakim = datetime.strptime(son_bakim, "%Y-%m-%d").date()

        ay_sayisi = parse_periyot(periyodu)
        if not ay_sayisi:
            continue

        # Son bakımdan ilerleyerek gelecek tarihleri hesapla
        gelecek_tarihler = []
        sonraki = son_bakim + relativedelta(months=ay_sayisi)

        while sonraki <= YIL_SONU:
            if sonraki >= BUGUN:
                gelecek_tarihler.append(sonraki)
            sonraki += relativedelta(months=ay_sayisi)

        if gelecek_tarihler:
            sonuc.append({
                "kodu": kodu,
                "tanimi": tanimi,
                "periyot_ay": ay_sayisi,
                "periyot_str": periyodu,
                "tarihler": gelecek_tarihler,
                "son_bakim": son_bakim,
            })

    # Sırala: kodu, sonra periyot
    sonuc.sort(key=lambda x: (x["kodu"], x["periyot_ay"]))
    return sonuc


# ─── Hafta Numarası Bazlı Gantt ─────────────────────────────────────────────
def hafta_listesi():
    """16 Nisan 2026'dan 31 Aralık 2026'ya kadar hafta başlangıç tarihlerini döndürür."""
    haftalar = []
    # İlk Pazartesiyi bul
    d = BUGUN
    # Haftanın Pazartesisine geri git
    d = d - timedelta(days=d.weekday())
    while d <= YIL_SONU:
        haftalar.append(d)
        d += timedelta(days=7)
    return haftalar


def tarih_haftaya(tarih, haftalar):
    """Bir tarihin hangi hafta sütununa denk geldiğini döndürür."""
    for i, h_start in enumerate(haftalar):
        h_end = h_start + timedelta(days=6)
        if h_start <= tarih <= h_end:
            return i
    return None


# ─── Excel Oluşturma ────────────────────────────────────────────────────────
def build_excel(veriler):
    wb = Workbook()

    # ═══════════════════════════════════════════
    # SAYFA 1: GANTT CHART
    # ═══════════════════════════════════════════
    ws = wb.active
    ws.title = "Gantt Takvim"
    ws.sheet_properties.tabColor = "3498DB"

    haftalar = hafta_listesi()
    n_hafta = len(haftalar)

    # Stiller
    baslik_font = Font(name="Segoe UI", bold=True, size=14, color="FFFFFF")
    baslik_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(name="Segoe UI", bold=True, size=8, color="FFFFFF")
    header_fill = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
    ay_header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    data_font = Font(name="Segoe UI", size=8)
    data_font_bold = Font(name="Segoe UI", size=8, bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=False)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="BDC3C7"),
        right=Side(style="thin", color="BDC3C7"),
        top=Side(style="thin", color="BDC3C7"),
        bottom=Side(style="thin", color="BDC3C7"),
    )
    zebra_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    bugun_fill = PatternFill(start_color="FFFDE7", end_color="FFFDE7", fill_type="solid")

    # Sabit sütunlar: #, Ekipman Kodu, Tanımı, Periyot, Son Bakım
    sabit_cols = 5
    col_hafta_start = sabit_cols + 1  # Hafta sütunları buradan başlar

    # ── Başlık Satırı (Row 1) ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=sabit_cols + n_hafta)
    c = ws.cell(row=1, column=1,
                value="2026 YILI PLANLI BAKIM GANTT TAKVİMİ  (16 Nisan – 31 Aralık)")
    c.font = baslik_font
    c.fill = baslik_fill
    c.alignment = center
    ws.row_dimensions[1].height = 28

    # ── Ay Satırı (Row 2) - Haftaları aylara göre grupla ──
    ay_gruplar = defaultdict(list)  # ay_no -> [hafta_idx, ...]
    for i, h in enumerate(haftalar):
        ay_gruplar[h.month].append(i)

    for ay_no, idxler in ay_gruplar.items():
        start_col = col_hafta_start + idxler[0]
        end_col = col_hafta_start + idxler[-1]
        if start_col != end_col:
            ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=end_col)
        c = ws.cell(row=2, column=start_col, value=AY_ISIMLERI[ay_no])
        c.font = Font(name="Segoe UI", bold=True, size=9, color="FFFFFF")
        c.fill = ay_header_fill
        c.alignment = center
        c.border = thin_border

    # Sabit sütun başlıkları da Row 2'ye
    for col_idx in range(1, sabit_cols + 1):
        ws.cell(row=2, column=col_idx).fill = ay_header_fill
    ws.row_dimensions[2].height = 20

    # ── Alt Başlık (Row 3) - Hafta tarihleri ──
    headers_sabit = ["#", "Ekipman Kodu", "Ekipman Tanımı", "Periyot", "Son Bakım"]
    col_widths_sabit = [5, 16, 45, 12, 12]

    for i, h in enumerate(headers_sabit):
        c = ws.cell(row=3, column=i + 1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
        ws.column_dimensions[get_column_letter(i + 1)].width = col_widths_sabit[i]

    for i, h_date in enumerate(haftalar):
        col = col_hafta_start + i
        c = ws.cell(row=3, column=col, value=h_date.strftime("%d.%m"))
        c.font = Font(name="Segoe UI", size=7, bold=True, color="FFFFFF")
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
        ws.column_dimensions[get_column_letter(col)].width = 4.5

    ws.row_dimensions[3].height = 18

    # Bugünün haftasını işaretle
    bugun_hafta_idx = tarih_haftaya(BUGUN, haftalar)

    # ── Veri Satırları ──
    row = 4
    for idx, v in enumerate(veriler, 1):
        is_zebra = idx % 2 == 0
        bg = zebra_fill if is_zebra else None

        # Sabit sütunlar
        periyot_ay = v["periyot_ay"]
        renk = PERIYOT_RENK.get(periyot_ay, "95A5A6")

        vals = [
            idx,
            v["kodu"],
            v["tanimi"],
            v["periyot_str"].replace("Periyodik: ", ""),
            v["son_bakim"].strftime("%d.%m.%Y") if v["son_bakim"] else "",
        ]

        for ci, val in enumerate(vals):
            c = ws.cell(row=row, column=ci + 1, value=val)
            c.font = data_font_bold if ci == 1 else data_font
            c.alignment = left if ci == 2 else center
            c.border = thin_border
            if bg:
                c.fill = bg

        # Periyot hücresini renklendir
        per_fill = PatternFill(start_color=PERIYOT_RENK_ACIK.get(periyot_ay, "F0F0F0"),
                               end_color=PERIYOT_RENK_ACIK.get(periyot_ay, "F0F0F0"),
                               fill_type="solid")
        ws.cell(row=row, column=4).fill = per_fill

        # Hafta sütunlarını doldur
        for hi in range(n_hafta):
            col = col_hafta_start + hi
            c = ws.cell(row=row, column=col)
            c.border = thin_border
            if bg:
                c.fill = bg
            # Bugünün haftasını vurgula
            if hi == bugun_hafta_idx:
                c.fill = bugun_fill

        # Bakım tarihlerini Gantt çubuğu olarak işaretle
        gantt_fill = PatternFill(start_color=renk, end_color=renk, fill_type="solid")
        for t in v["tarihler"]:
            hi = tarih_haftaya(t, haftalar)
            if hi is not None:
                col = col_hafta_start + hi
                c = ws.cell(row=row, column=col)
                c.value = t.day  # Günü yaz
                c.font = Font(name="Segoe UI", size=7, bold=True, color="FFFFFF")
                c.fill = gantt_fill
                c.alignment = center

        ws.row_dimensions[row].height = 16
        row += 1

    # ── Bugün İşaretçisi (dikey çizgi efekti) ──
    if bugun_hafta_idx is not None:
        bugun_col = col_hafta_start + bugun_hafta_idx
        for r in range(3, row):
            c = ws.cell(row=r, column=bugun_col)
            if not c.fill or c.fill.start_color.index in ("00000000", "F8F9FA"):
                c.fill = bugun_fill

    # ── Lejant Satırları ──
    row += 1
    ws.cell(row=row, column=1, value="LEJANT:").font = Font(name="Segoe UI", bold=True, size=9)
    row += 1

    lejant = [
        (1, "1 Aylık Bakım"),
        (3, "3 Aylık Bakım"),
        (6, "6 Aylık Bakım"),
        (12, "Yıllık Bakım"),
    ]
    for ay, label in lejant:
        renk_hex = PERIYOT_RENK[ay]
        c = ws.cell(row=row, column=2)
        c.fill = PatternFill(start_color=renk_hex, end_color=renk_hex, fill_type="solid")
        c.border = thin_border
        ws.cell(row=row, column=3, value=label).font = data_font
        row += 1

    c = ws.cell(row=row, column=2)
    c.fill = bugun_fill
    c.border = thin_border
    ws.cell(row=row, column=3, value="Bugün (16 Nisan 2026)").font = data_font
    row += 2

    # İstatistik
    toplam_bakim = sum(len(v["tarihler"]) for v in veriler)
    ws.cell(row=row, column=1,
            value=f"Toplam: {len(veriler)} ekipman grubu, {toplam_bakim} planlanan bakım").font = \
        Font(name="Segoe UI", size=9, italic=True, color="7F8C8D")

    # Freeze panes
    ws.freeze_panes = ws.cell(row=4, column=col_hafta_start)

    # Auto-filter
    ws.auto_filter.ref = f"A3:{get_column_letter(sabit_cols + n_hafta)}{3 + len(veriler)}"

    # ═══════════════════════════════════════════
    # SAYFA 2: AYLIK ÖZET
    # ═══════════════════════════════════════════
    ws2 = wb.create_sheet("Aylık Özet")
    ws2.sheet_properties.tabColor = "2ECC71"

    # Aylık sayıları hesapla
    aylik = defaultdict(lambda: defaultdict(int))
    for v in veriler:
        for t in v["tarihler"]:
            aylik[t.month][v["periyot_ay"]] += 1

    # Başlık
    ws2.merge_cells("A1:F1")
    c = ws2.cell(row=1, column=1, value="2026 AYLIK PLANLI BAKIM ÖZETİ (Nisan – Aralık)")
    c.font = baslik_font
    c.fill = baslik_fill
    c.alignment = center

    # Tablo başlıkları
    ozet_headers = ["Ay", "1 Aylık", "3 Aylık", "6 Aylık", "Yıllık", "TOPLAM"]
    for i, h in enumerate(ozet_headers):
        c = ws2.cell(row=3, column=i + 1, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border

    ws2.column_dimensions["A"].width = 14
    for col_l in ["B", "C", "D", "E", "F"]:
        ws2.column_dimensions[col_l].width = 12

    genel_toplam = 0
    periyot_toplamlar = defaultdict(int)
    r = 4
    for ay_no in range(4, 13):
        is_zebra = (ay_no - 4) % 2 == 1
        ay_toplam = 0
        ws2.cell(row=r, column=1, value=AY_ISIMLERI[ay_no]).font = data_font_bold
        ws2.cell(row=r, column=1).alignment = center
        ws2.cell(row=r, column=1).border = thin_border
        if is_zebra:
            ws2.cell(row=r, column=1).fill = zebra_fill

        for pi, pay in enumerate([1, 3, 6, 12]):
            val = aylik[ay_no].get(pay, 0)
            ay_toplam += val
            periyot_toplamlar[pay] += val
            c = ws2.cell(row=r, column=pi + 2, value=val if val > 0 else "-")
            c.font = data_font
            c.alignment = center
            c.border = thin_border
            if val > 0:
                c.fill = PatternFill(start_color=PERIYOT_RENK_ACIK[pay],
                                     end_color=PERIYOT_RENK_ACIK[pay],
                                     fill_type="solid")
            elif is_zebra:
                c.fill = zebra_fill

        genel_toplam += ay_toplam
        c = ws2.cell(row=r, column=6, value=ay_toplam)
        c.font = data_font_bold
        c.alignment = center
        c.border = thin_border
        c.fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
        r += 1

    # Toplam satırı
    ws2.cell(row=r, column=1, value="TOPLAM").font = Font(name="Segoe UI", bold=True, size=9, color="FFFFFF")
    ws2.cell(row=r, column=1).fill = header_fill
    ws2.cell(row=r, column=1).alignment = center
    ws2.cell(row=r, column=1).border = thin_border
    for pi, pay in enumerate([1, 3, 6, 12]):
        c = ws2.cell(row=r, column=pi + 2, value=periyot_toplamlar[pay])
        c.font = Font(name="Segoe UI", bold=True, size=9, color="FFFFFF")
        c.fill = header_fill
        c.alignment = center
        c.border = thin_border
    c = ws2.cell(row=r, column=6, value=genel_toplam)
    c.font = Font(name="Segoe UI", bold=True, size=9, color="FFFFFF")
    c.fill = PatternFill(start_color="1ABC9C", end_color="1ABC9C", fill_type="solid")
    c.alignment = center
    c.border = thin_border

    # ═══════════════════════════════════════════
    # KAYDET
    # ═══════════════════════════════════════════
    wb.save(OUTPUT_FILE)
    return OUTPUT_FILE, len(veriler), toplam_bakim


# ─── Ana ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    print("Veritabanından son bakım tarihleri çekiliyor...")
    rows = fetch_last_maintenance()
    print(f"  {len(rows)} ekipman+periyot grubu bulundu")

    print("Gelecek bakım tarihleri hesaplanıyor (16.04.2026 – 31.12.2026)...")
    veriler = hesapla_gelecek_bakimlar(rows)
    print(f"  {len(veriler)} grup için gelecek bakım planlandı")
    print(f"  Toplam {sum(len(v['tarihler']) for v in veriler)} bakım işlemi")

    print("Excel Gantt takvimi oluşturuluyor...")
    path, grup, toplam = build_excel(veriler)
    print(f"Excel oluşturuldu: {path}")
    print(f"  → {grup} ekipman grubu, {toplam} planlanan bakım")
